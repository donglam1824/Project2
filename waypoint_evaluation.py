import pandas as pd
import os

def read_map_from_file(filename):
    grid = []
    start = None
    goal = None
    
    with open(filename, 'r') as file:
        for y, line in enumerate(file):
            line = line.strip()
            if not line:
                continue
                
            row = []
            parts = line.split()
            
            for x, val in enumerate(parts):
                if val == '*':
                    start = (y, x)
                    row.append(0)  # Giả sử vị trí bắt đầu là 0 (có thể đi qua)
                elif val == '#':
                    goal = (y, x)
                    row.append(0)  # Giả sử vị trí đích là 0 (có thể đi qua)
                else:
                    row.append(int(val))
            
            grid.append(row)
    
    if start is None or goal is None:
        raise ValueError("Map must contain both start '*' and goal '#'")
    
    return grid, start, goal


def read_waypoints_from_file(filename):
    waypoints = []
    with open(filename, 'r') as file:
        for line in file:
            line = line.strip()
            if line:
                waypoint = tuple(map(int, line.strip('()').split(',')))
                waypoints.append(waypoint)
    return waypoints

def is_valid_path(waypoints, map_grid):
    max_y = len(map_grid)
    max_x = len(map_grid[0]) if max_y > 0 else 0

    for i, (y, x) in enumerate(waypoints):
        # 1. Kiểm tra trong giới hạn bản đồ
        if not (0 <= y < max_y and 0 <= x < max_x):
            print(f"❌ Waypoint {i} ({y}, {x}) nằm ngoài bản đồ.")
            return False

        # 2. Kiểm tra ô có thể đi
        if map_grid[y][x] != 0:
            print(f"❌ Waypoint {i} ({y}, {x}) không nằm trên ô có thể đi (giá trị: {map_grid[y][x]}).")
            return False

        # 3. Kiểm tra liền kề (bao gồm đi chéo)
        if i > 0:
            y_prev, x_prev = waypoints[i - 1]
            dy = abs(y - y_prev)
            dx = abs(x - x_prev)

            if max(dx, dy) != 1:
                print(f"❌ Waypoint {i} ({y}, {x}) không liền kề waypoint trước đó ({y_prev}, {x_prev}).")
                return False

    return True 

def evaluate_waypoints(waypoints):
    total_distance = 0
    turns = 0
    prev_direction = None

    for i in range(1, len(waypoints)):
        x1, y1 = waypoints[i - 1]
        x2, y2 = waypoints[i]

        dx = x2 - x1
        dy = y2 - y1

        total_distance += 1

        if dx != 0:
            direction = 'horizontal'
        elif dy != 0:
            direction = 'vertical'
        else:
            direction = prev_direction

        if prev_direction and direction != prev_direction:
            turns += 1

        prev_direction = direction

    return total_distance, turns

def run_waypoint_evaluation_from_file(map_grid, filename):
    waypoints = read_waypoints_from_file(filename)
    r, tau = evaluate_waypoints(waypoints)
    map_height = len(map_grid)
    map_width = len(map_grid[0]) if map_height > 0 else 0
    map_size = f"{map_width}x{map_height}"

    if not is_valid_path(waypoints, map_grid):
        print("Không được đi qua 1")
        return

    print(f"Waypoints: {waypoints}")
    print(f"Distance (steps): {r}, Turns: {tau}")

    visited_cells = set(waypoints)

    total_walkable_cells = sum(row.count(0) for row in map_grid)
    coverage_cells = len(visited_cells)
    coverage_ratio = coverage_cells / total_walkable_cells if total_walkable_cells > 0 else 0

    print(f"Length: {r}")
    print(f"Coverage Ratio: {coverage_ratio:.2%}")

    # if r <= EVAL_THRESHOLD and tau <= EVAL_THRESHOLD:
    #     print("Waypoints accepted.")
    # else:
    #     print("Waypoints rejected based on evaluation criteria.")
    
    # Đường dẫn file Excel
    excel_file = "bfs.xlsx"
    
    # Chuẩn bị dữ liệu mới
    new_data = {
        'Kích thước bản đồ': map_size,
        'Độ dài đường đi': [r],
        'Số lần rẽ': [tau],
        'Tổng số ô có thể đi': [total_walkable_cells],
        'Số waypoint': [len(waypoints)],
        'Tỉ lệ bao phủ': [coverage_ratio],
        'LLMs': filename[9:-4]
    }
    new_df = pd.DataFrame(new_data)
    
    # Chuẩn bị dữ liệu waypoints mới
    new_waypoints = pd.DataFrame(waypoints, columns=['Y', 'X'])
    
    try:
        # Kiểm tra xem file Excel đã tồn tại chưa
        if os.path.exists(excel_file):
            # Sử dụng openpyxl trực tiếp để bảo toàn công thức
            from openpyxl import load_workbook
            
            # Đọc workbook hiện có
            wb = load_workbook(excel_file)
            
            # Xử lý sheet "Thông số"
            if "Thông số" in wb.sheetnames:
                # Đọc sheet hiện có để xác định run_id và thêm dữ liệu mới
                stats_sheet = wb["Thông số"]
                
                # Đếm số dòng hiện có (bỏ qua header)
                row_count = stats_sheet.max_row - 1
                
                # Thêm dữ liệu mới vào sheet
                row_idx = stats_sheet.max_row + 1
                for col_idx, col_name in enumerate(new_data.keys(), start=1):
                    stats_sheet.cell(row=row_idx, column=col_idx).value = new_data[col_name][0]
            else:
                # Nếu sheet chưa tồn tại, tạo mới
                stats_sheet = wb.create_sheet("Thông số")
                
                # Thêm header
                for col_idx, col_name in enumerate(new_data.keys(), start=1):
                    stats_sheet.cell(row=1, column=col_idx).value = col_name
                
                # Thêm dữ liệu
                for col_idx, col_name in enumerate(new_data.keys(), start=1):
                    stats_sheet.cell(row=2, column=col_idx).value = new_data[col_name][0]
                
                row_count = 1  # Chỉ có 1 row (không tính header)
            
            # Xử lý sheet "Waypoints"
            if "Waypoints" in wb.sheetnames:
                waypoints_sheet = wb["Waypoints"]
                
                # Xác định run_id từ số dòng trong sheet "Thông số"
                run_id = row_count
                
                # Thêm dữ liệu waypoints mới vào sheet
                start_row = waypoints_sheet.max_row + 1
                for i, (y, x) in enumerate(waypoints):
                    waypoints_sheet.cell(row=start_row+i, column=1).value = y
                    waypoints_sheet.cell(row=start_row+i, column=2).value = x
                    waypoints_sheet.cell(row=start_row+i, column=3).value = run_id
            else:
                # Nếu sheet chưa tồn tại, tạo mới
                waypoints_sheet = wb.create_sheet("Waypoints")
                
                # Thêm header
                waypoints_sheet.cell(row=1, column=1).value = "Y"
                waypoints_sheet.cell(row=1, column=2).value = "X"
                waypoints_sheet.cell(row=1, column=3).value = "Run ID"
                
                # Thêm dữ liệu
                run_id = row_count
                for i, (y, x) in enumerate(waypoints):
                    waypoints_sheet.cell(row=i+2, column=1).value = y
                    waypoints_sheet.cell(row=i+2, column=2).value = x
                    waypoints_sheet.cell(row=i+2, column=3).value = run_id
            
            # Lưu workbook
            wb.save(excel_file)
            print(f"Data written and preserved all sheets and formulas: {excel_file}")
        else:
            # Nếu file chưa tồn tại, sử dụng pandas để tạo mới
            run_id = 1
            new_waypoints['Run ID'] = run_id
            
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                new_df.to_excel(writer, sheet_name='Thông số', index=False)
                new_waypoints.to_excel(writer, sheet_name='Waypoints', index=False)
                
            print(f"Created new file: {excel_file}")
            
    except PermissionError:
        # Nếu gặp lỗi quyền truy cập, thử lưu với tên file khác
        alt_excel_file = "waypoint_report_new.xlsx"
        with pd.ExcelWriter(alt_excel_file, engine='openpyxl') as writer:
            new_df.to_excel(writer, sheet_name='Thông số', index=False)
            new_waypoints.to_excel(writer, sheet_name='Waypoints', index=False)
        print(f"Can not write {excel_file}. Instead: {alt_excel_file}")


if __name__ == "__main__":
    map_grid, start_position, target_position = read_map_from_file('input_map.txt')
    print(f"Start: {start_position}, Goal: {target_position}")
    run_waypoint_evaluation_from_file(map_grid, 'waypoint_gpt.txt')
    # run_waypoint_evaluation_from_file(map_grid, 'waypoint_gemini.txt')