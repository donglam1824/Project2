import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import ListedColormap
import matplotlib.patches as mpatches
from datetime import datetime
import os
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class WaypointAnalyzer:
    def __init__(self, map_file, waypoint_file):
        self.map_file = map_file
        self.waypoint_file = waypoint_file
        self.map_grid = None
        self.waypoints = []
        self.load_map()
        self.load_waypoints()
    
    def load_map(self):
        """Đọc bản đồ từ file"""
        with open(self.map_file, 'r') as f:
            lines = f.readlines()
        
        self.map_grid = []
        for line in lines:
            row = [int(x) for x in line.strip().split()]
            self.map_grid.append(row)
        
        self.map_grid = np.array(self.map_grid)
        print(f"Map loaded: {self.map_grid.shape[0]}x{self.map_grid.shape[1]}")
    
    def load_waypoints(self):
        """Đọc waypoints từ file"""
        with open(self.waypoint_file, 'r') as f:
            for line in f:
                if line.strip():
                    # Parse (x, y) format
                    coords = line.strip().replace('(', '').replace(')', '').split(',')
                    x, y = int(coords[0].strip()), int(coords[1].strip())
                    self.waypoints.append((x, y))
        
        print(f"Loaded {len(self.waypoints)} waypoints")
    
    def calculate_path_distance(self):
        """Tính tổng độ dài đường đi"""
        if len(self.waypoints) < 2:
            return 0
        
        total_distance = 0
        for i in range(1, len(self.waypoints)):
            x1, y1 = self.waypoints[i-1]
            x2, y2 = self.waypoints[i]
            # Tính khoảng cách Euclidean
            distance = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)
            total_distance += distance
        
        return total_distance
    
    def calculate_coverage_area(self):
        """Tính diện tích không gian đi được (số ô đã đi qua)"""
        visited_cells = set()
        
        for x, y in self.waypoints:
            # Kiểm tra xem tọa độ có hợp lệ không
            if (0 <= x < self.map_grid.shape[0] and 
                0 <= y < self.map_grid.shape[1] and 
                self.map_grid[x, y] == 0):  # 0 là không gian trống
                visited_cells.add((x, y))
        
        # Tính tổng số ô trống có thể đi được
        total_free_cells = np.sum(self.map_grid == 0)
        
        coverage_ratio = len(visited_cells) / total_free_cells if total_free_cells > 0 else 0
        
        return len(visited_cells), total_free_cells, coverage_ratio
    
    def calculate_direction_changes(self):
        """Tính số lần quay đầu (thay đổi hướng)"""
        if len(self.waypoints) < 3:
            return 0
        
        direction_changes = 0
        
        for i in range(2, len(self.waypoints)):
            # Lấy 3 điểm liên tiếp
            p1 = self.waypoints[i-2]
            p2 = self.waypoints[i-1]
            p3 = self.waypoints[i]
            
            # Tính vector hướng
            vec1 = (p2[0] - p1[0], p2[1] - p1[1])
            vec2 = (p3[0] - p2[0], p3[1] - p2[1])
            
            # Kiểm tra xem có thay đổi hướng không
            if vec1 != vec2 and vec1 != (0, 0) and vec2 != (0, 0):
                # Tính góc giữa 2 vector
                dot_product = vec1[0] * vec2[0] + vec1[1] * vec2[1]
                mag1 = math.sqrt(vec1[0]**2 + vec1[1]**2)
                mag2 = math.sqrt(vec2[0]**2 + vec2[1]**2)
                
                if mag1 > 0 and mag2 > 0:
                    cos_angle = dot_product / (mag1 * mag2)
                    cos_angle = max(-1, min(1, cos_angle))  # Clamp to [-1, 1]
                    angle = math.acos(cos_angle)
                    
                    # Nếu góc thay đổi > 45 độ thì coi là quay đầu
                    if angle > math.pi/4:
                        direction_changes += 1
        
        return direction_changes
    
    def analyze_efficiency(self):
        """Phân tích hiệu quả của đường đi"""
        path_distance = self.calculate_path_distance()
        visited_cells, total_free_cells, coverage_ratio = self.calculate_coverage_area()
        direction_changes = self.calculate_direction_changes()
        
        # Tính hiệu quả (coverage per unit distance)
        efficiency = coverage_ratio / path_distance if path_distance > 0 else 0
        
        return {
            'path_distance': path_distance,
            'visited_cells': visited_cells,
            'total_free_cells': total_free_cells,
            'coverage_ratio': coverage_ratio,
            'direction_changes': direction_changes,
            'efficiency': efficiency
        }
    
    def save_to_excel(self, filename='apartment.xlsx'):
        """Lưu kết quả phân tích vào file Excel (append mode)"""
        results = self.analyze_efficiency()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Kiểm tra xem file đã tồn tại chưa
        file_exists = os.path.exists(filename)
        
        if file_exists:
            # Đọc file existing để append dữ liệu
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename)
                
                # Tạo hoặc lấy sheet Summary
                if "Analysis History" in wb.sheetnames:
                    ws_history = wb["Analysis History"]
                else:
                    ws_history = wb.create_sheet("Analysis History")
                    # Thêm header nếu sheet mới
                    headers = ["Timestamp", "Map File", "Waypoint File", "Total Waypoints", 
                              "Path Distance", "Coverage Ratio", "Direction Changes", 
                              "Efficiency", "Visited Cells", "Total Free Cells"]
                    for col, header in enumerate(headers, 1):
                        cell = ws_history.cell(row=1, column=col, value=header)
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                           top=Side(style='thin'), bottom=Side(style='thin'))
                
                # Tìm dòng trống tiếp theo
                next_row = ws_history.max_row + 1
                
            except Exception as e:
                print(f"Error loading existing file: {e}")
                # Nếu có lỗi, tạo file mới
                wb = Workbook()
                ws_history = wb.active
                ws_history.title = "Analysis History"
                next_row = 1
        else:
            # Tạo workbook mới
            wb = Workbook()
            ws_history = wb.active
            ws_history.title = "Analysis History"
            next_row = 1
        
        # Nếu là dòng đầu tiên, thêm header
        if next_row == 1:
            headers = ["Timestamp", "Map File", "Waypoint File", "Total Waypoints", 
                      "Path Distance", "Coverage Ratio", "Direction Changes", 
                      "Efficiency", "Visited Cells", "Total Free Cells"]
            for col, header in enumerate(headers, 1):
                cell = ws_history.cell(row=next_row, column=col, value=header)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center', vertical='center')
            next_row += 1
        
        # Thêm dữ liệu mới
        new_data = [
            current_time,
            self.map_file,
            self.waypoint_file,
            len(self.waypoints),
            round(results['path_distance'], 2),
            f"{results['coverage_ratio']:.2%}",
            results['direction_changes'],
            round(results['efficiency'], 4),
            results['visited_cells'],
            results['total_free_cells']
        ]
        
        for col, value in enumerate(new_data, 1):
            cell = ws_history.cell(row=next_row, column=col, value=value)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Tạo/cập nhật sheet chi tiết cho lần chạy hiện tại
        sheet_name = f"Detail_{datetime.now().strftime('%m%d_%H%M%S')}"
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)
        
        # Định dạng màu sắc và font cho sheet detail
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        metric_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        metric_font = Font(bold=True, size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Header
        ws['A1'] = "WAYPOINT ANALYSIS REPORT"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_align
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Analysis Date: {current_time}"
        ws['A2'].font = Font(size=10)
        ws.merge_cells('A2:D2')
        
        # Thông tin cơ bản
        row = 4
        basic_info = [
            ("Map Information", ""),
            ("Map Size", f"{self.map_grid.shape[0]} x {self.map_grid.shape[1]}"),
            ("Total Waypoints", len(self.waypoints)),
            ("Map File", self.map_file),
            ("Waypoint File", self.waypoint_file)
        ]
        
        for info in basic_info:
            ws[f'A{row}'] = info[0]
            ws[f'B{row}'] = info[1]
            if info[0] == "Map Information":
                ws[f'A{row}'].fill = header_fill
                ws[f'A{row}'].font = header_font
                ws[f'B{row}'].fill = header_fill
                ws[f'B{row}'].font = header_font
            else:
                ws[f'A{row}'].font = metric_font
            row += 1
        
        # Metrics chính
        row += 1
        metrics_data = [
            ("Path Metrics", ""),
            ("Total Path Distance", f"{results['path_distance']:.2f} units"),
            ("Direction Changes", results['direction_changes']),
            ("", ""),
            ("Coverage Metrics", ""),
            ("Visited Cells", results['visited_cells']),
            ("Total Free Cells", results['total_free_cells']),
            ("Coverage Ratio", f"{results['coverage_ratio']:.2%}"),
            ("", ""),
            ("Efficiency Metrics", ""),
            ("Efficiency (Coverage/Distance)", f"{results['efficiency']:.4f}"),
            ("Average Distance per Cell", f"{results['path_distance']/results['visited_cells']:.2f}" if results['visited_cells'] > 0 else "N/A")
        ]
        
        for metric in metrics_data:
            ws[f'A{row}'] = metric[0]
            ws[f'B{row}'] = metric[1]
            if metric[0] in ["Path Metrics", "Coverage Metrics", "Efficiency Metrics"]:
                ws[f'A{row}'].fill = header_fill
                ws[f'A{row}'].font = header_font
                ws[f'B{row}'].fill = header_fill
                ws[f'B{row}'].font = header_font
            elif metric[0] != "":
                ws[f'A{row}'].font = metric_font
            row += 1
        
        # Tạo bảng waypoints
        row += 1
        ws[f'A{row}'] = "Waypoint Details"
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ["Index", "X Coordinate", "Y Coordinate", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = metric_fill
            cell.font = metric_font
            cell.border = border
            cell.alignment = center_align
        
        # Thêm dữ liệu waypoints
        for idx, (x, y) in enumerate(self.waypoints):
            row += 1
            # Kiểm tra trạng thái của waypoint
            status = "Valid" if (0 <= x < self.map_grid.shape[0] and 
                               0 <= y < self.map_grid.shape[1] and 
                               self.map_grid[x, y] == 0) else "Invalid"
            
            data = [idx + 1, x, y, status]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = center_align
                if status == "Invalid" and col > 1:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Tạo sheet thống kê tổng hợp nếu chưa có
        if "Statistics Summary" not in wb.sheetnames:
            ws2 = wb.create_sheet("Statistics Summary")
            
            summary_data = [
                ["Metric", "Current Value", "Unit", "Description"],
                ["Total Waypoints", len(self.waypoints), "points", "Number of waypoints in the path"],
                ["Path Distance", f"{results['path_distance']:.2f}", "units", "Total Euclidean distance traveled"],
                ["Coverage Ratio", f"{results['coverage_ratio']:.2%}", "%", "Percentage of free space visited"],
                ["Direction Changes", results['direction_changes'], "changes", "Number of significant direction changes"],
                ["Efficiency Score", f"{results['efficiency']:.4f}", "ratio", "Coverage per unit distance"],
                ["Map Utilization", f"{results['visited_cells']}/{results['total_free_cells']}", "cells", "Visited cells / Total free cells"]
            ]
            
            for row_idx, data in enumerate(summary_data, 1):
                for col_idx, value in enumerate(data, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # Header row
                        cell.fill = header_fill
                        cell.font = header_font
                    cell.border = border
                    cell.alignment = center_align
        else:
            ws2 = wb["Statistics Summary"]
            # Cập nhật giá trị hiện tại
            ws2['B2'] = len(self.waypoints)
            ws2['B3'] = f"{results['path_distance']:.2f}"
            ws2['B4'] = f"{results['coverage_ratio']:.2%}"
            ws2['B5'] = results['direction_changes']
            ws2['B6'] = f"{results['efficiency']:.4f}"
            ws2['B7'] = f"{results['visited_cells']}/{results['total_free_cells']}"
        
        # Điều chỉnh độ rộng cột cơ bản (tránh lỗi MergedCell)
        from openpyxl.utils import get_column_letter
        
        for sheet in wb.worksheets:
            # Đặt độ rộng cố định cho các cột chính để tránh lỗi
            try:
                if sheet.max_column >= 1:
                    sheet.column_dimensions['A'].width = 20  # Timestamp/Metric
                if sheet.max_column >= 2:
                    sheet.column_dimensions['B'].width = 15  # Value
                if sheet.max_column >= 3:
                    sheet.column_dimensions['C'].width = 15  # Unit/File
                if sheet.max_column >= 4:
                    sheet.column_dimensions['D'].width = 25  # Description
                if sheet.max_column >= 5:
                    for col_num in range(5, min(sheet.max_column + 1, 11)):
                        sheet.column_dimensions[get_column_letter(col_num)].width = 12
            except Exception as e:
                print(f"Warning: Could not adjust column widths for {sheet.title}: {e}")
                pass
        
        # Lưu file
        wb.save(filename)
        print(f"Analysis results appended to {filename}")
        print(f"New analysis saved in sheet: {sheet_name}")
        return filename
    
    def visualize_path(self, save_image=True):
        """Vẽ bản đồ và đường đi"""
        # Tạo thư mục image nếu chưa tồn tại
        if not os.path.exists('image'):
            os.makedirs('image')
        
        plt.figure(figsize=(15, 12))
        ax = plt.gca()
        
        # Sử dụng màu sắc như code gốc: trắng (trống), đỏ (tường)
        colors = ['#FFFFFF', '#F44336', '#4CAF50']
        cmap = ListedColormap(colors)
        
        # Vẽ bản đồ gốc không thay đổi màu
        plt.pcolormesh(self.map_grid, cmap=cmap, edgecolors='k', linewidth=1.5)
        
        # Vẽ đường đi
        if self.waypoints:
            x_coords = [coord[1] for coord in self.waypoints]  # y coordinate for plotting
            y_coords = [coord[0] for coord in self.waypoints]  # x coordinate for plotting
            x_plot = [x + 0.5 for x in x_coords]
            y_plot = [y + 0.5 for y in y_coords]
            
            # Vẽ đường đi
            plt.plot(x_plot, y_plot, color='#2196F3', linewidth=2, alpha=0.8, label='Robot Path')
            
            # Đánh dấu điểm bắt đầu và kết thúc
            plt.plot(x_plot[0], y_plot[0], 'go', markersize=10, label='Start')
            plt.plot(x_plot[-1], y_plot[-1], 'ro', markersize=10, label='End')
            
            # Hiển thị số thứ tự một số điểm
            step = 10  
            for i in range(0, len(x_plot), step):
                plt.text(x_plot[i], y_plot[i], str(i), fontsize=8, 
                        ha='center', va='center', 
                        bbox=dict(boxstyle='round,pad=0.2', facecolor='white', alpha=0.7))
        
        plt.gca().invert_yaxis()
        plt.title('Robot Path Visualization', fontsize=16, fontweight='bold')
        
        # Tạo legend
        legend_elements = [
            mpatches.Patch(color='#FFFFFF', label='Free Space'),
            mpatches.Patch(color='#F44336', label='Wall'),
            mpatches.Patch(color='#4CAF50', label='Visited'),
            plt.Line2D([0], [0], color='#2196F3', linewidth=2, label='Robot Path'),
            plt.Line2D([0], [0], marker='o', color='g', linewidth=0, markersize=8, label='Start'),
            plt.Line2D([0], [0], marker='o', color='r', linewidth=0, markersize=8, label='End')
        ]
        plt.legend(handles=legend_elements, loc='upper right', bbox_to_anchor=(1.15, 1))
        
        plt.tight_layout()
        
        if save_image:
            current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_image_path = os.path.join('image', f'path_analysis_{current_time}.png')
            plt.savefig(output_image_path, dpi=300, bbox_inches='tight')
            print(f"Analysis image saved to {output_image_path}")
        
        plt.show()
    
    def print_analysis_report(self):
        """In báo cáo phân tích chi tiết"""
        results = self.analyze_efficiency()
        
        print("="*60)
        print("           WAYPOINT ANALYSIS REPORT")
        print("="*60)
        print(f"Map Size: {self.map_grid.shape[0]} x {self.map_grid.shape[1]}")
        print(f"Total Waypoints: {len(self.waypoints)}")
        print("-"*60)
        print("PATH METRICS:")
        print(f"  • Total Path Distance: {results['path_distance']:.2f} units")
        print(f"  • Direction Changes: {results['direction_changes']}")
        print("-"*60)
        print("COVERAGE METRICS:")
        print(f"  • Visited Cells: {results['visited_cells']}")
        print(f"  • Total Free Cells: {results['total_free_cells']}")
        print(f"  • Coverage Ratio: {results['coverage_ratio']:.2%}")
        print("-"*60)
        print("EFFICIENCY METRICS:")
        print(f"  • Efficiency (Coverage/Distance): {results['efficiency']:.4f}")
        print(f"  • Average Distance per Cell: {results['path_distance']/results['visited_cells']:.2f}" if results['visited_cells'] > 0 else "  • Average Distance per Cell: N/A")
        print("="*60)
        
        return results

# Sử dụng class
if __name__ == "__main__":
    # Khởi tạo analyzer
    analyzer = WaypointAnalyzer('apartment_map.txt', 'waypoint_gpt.txt')
    
    # Thực hiện phân tích và in báo cáo
    results = analyzer.print_analysis_report()
    
    # Lưu kết quả vào Excel
    excel_file = analyzer.save_to_excel('apartment.xlsx')
    
    # Vẽ và lưu hình ảnh
    analyzer.visualize_path(save_image=True)
    
    print(f"\nAll analysis completed! Results saved to {excel_file}")